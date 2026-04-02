"""
Bile Acid Analysis Pipeline - Streamlit Application
====================================================

Run with: streamlit run app.py
"""

import gc
import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from io import BytesIO
from pathlib import Path
import sys
import tempfile
import zipfile
from datetime import datetime

sys.path.insert(0, str(Path(__file__).parent))

from config.bile_acid_species import (
    BILE_ACID_PANEL, ANALYSIS_GROUPS, ANALYSIS_GROUP_DISPLAY_NAMES,
    get_glycine_conjugated, get_taurine_conjugated,
    get_primary, get_secondary, get_conjugated, get_unconjugated,
    get_keto_derivatives, get_iso_forms, get_nor_bile_acids,
    get_12alpha_hydroxylated, get_non12alpha_hydroxylated, get_sulfated
)
from modules.data_processing import BileAcidDataProcessor, ProcessedData, validate_data_quality
from modules.statistical_tests import StatisticalAnalyzer, format_analysis_report, format_twoway_apa, format_threeway_apa
from modules.visualization import BileAcidVisualizer
from modules.report_generation import (
    ExcelReportGenerator, SignificancePlotter,
    ComprehensiveAnalysisResults, format_apa_statistics,
    get_significant_differences_summary, get_twoway_differences_summary,
    get_threeway_differences_summary
)

st.set_page_config(page_title="Bile Acid Analysis Pipeline", page_icon="🧬", layout="wide")


def df_to_excel_bytes(df, number_format='0.00'):
    """Convert a DataFrame to Excel bytes with numeric cell formatting."""
    from openpyxl.utils import get_column_letter
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Sheet1', index=False)
        ws = writer.sheets['Sheet1']
        for col_idx, col_name in enumerate(df.columns, 1):
            if pd.api.types.is_numeric_dtype(df[col_name]):
                col_letter = get_column_letter(col_idx)
                for row_idx in range(2, len(df) + 2):
                    ws[f'{col_letter}{row_idx}'].number_format = number_format
    buf.seek(0)
    return buf.getvalue()


def init_session_state():
    """Initialize session state variables."""
    defaults = {
        'processed_data': None,
        'analysis_results': ComprehensiveAnalysisResults(),
        'figures': {},
        'report_generator': None,
        'stats_computed': False,
        'stats_sections_done': set(),  # Track which stat sections are computed
        'last_file': None,
        'last_settings': None,  # Track settings that affect data/stats
        'zip_data': None,  # Cached ZIP bytes for download
        'zip_figure_count': 0,  # Number of figures in cached ZIP
        'selected_sheet': None  # User-selected data sheet
    }
    for key, val in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = val


def get_data_affecting_settings(settings):
    """Extract settings that affect data processing and statistics."""
    return {
        'lod_handling': settings['lod_handling'],
        'lod_threshold': settings.get('lod_threshold', 50),
        'alpha': settings['alpha']
    }


def check_settings_changed(settings):
    """Check if data-affecting settings have changed, reset caches if so."""
    current = get_data_affecting_settings(settings)
    
    if st.session_state.last_settings != current:
        # Settings changed - need to reprocess data and recompute stats
        st.session_state.stats_computed = False
        st.session_state.stats_sections_done = set()
        st.session_state.report_generator = None
        st.session_state.figures = {}
        st.session_state.zip_data = None
        st.session_state.zip_figure_count = 0
        st.session_state.last_settings = current
        return True
    return False


def fig_to_bytes(fig, format='png', dpi=300):
    """Convert matplotlib figure to bytes."""
    buf = BytesIO()
    fig.savefig(buf, format=format, dpi=dpi, bbox_inches='tight')
    buf.seek(0)
    return buf.getvalue()


def store_figure(fig, name):
    """Store a figure as PNG+PDF bytes in session state (not the raw object)."""
    if 'figures' not in st.session_state:
        st.session_state.figures = {}
    st.session_state.figures[name] = {
        'png': fig_to_bytes(fig, 'png', dpi=300),
        'pdf': fig_to_bytes(fig, 'pdf'),
    }


def _ensure_report_generator(processed, settings):
    """Create ExcelReportGenerator if not already cached."""
    if st.session_state.report_generator is not None:
        return st.session_state.report_generator

    group_col = processed.structure.group_col
    factors = getattr(processed.structure, 'factors', {})
    n_factors = getattr(processed.structure, 'n_factors', 0)

    report_gen = ExcelReportGenerator(
        data=processed.sample_data,
        group_col=group_col,
        bile_acid_cols=processed.structure.bile_acid_cols,
        totals=processed.totals,
        ratios=processed.ratios,
        percentages=processed.percentages,
        alpha=settings['alpha'],
        factors=factors,
        n_factors=n_factors,
        analyte_lod_counts=getattr(processed.structure, 'analyte_lod_counts', {}),
        analyte_lod_rows=getattr(processed.structure, 'analyte_lod_rows', {}),
        n_samples=len(processed.sample_data),
        lod_threshold=settings.get('lod_threshold', 50),
    )
    st.session_state.report_generator = report_gen
    return report_gen


def ensure_section_computed(processed, settings, section):
    """Compute stats for a single section if not already done."""
    if section in st.session_state.stats_sections_done:
        return st.session_state.analysis_results

    group_col = processed.structure.group_col
    if not group_col:
        return st.session_state.analysis_results

    report_gen = _ensure_report_generator(processed, settings)
    results = report_gen.run_section(section)
    st.session_state.analysis_results = results
    st.session_state.stats_sections_done.add(section)
    return results


def compute_all_statistics(processed, settings):
    """Compute all statistics at once (used for one-way and two-way designs)."""
    if st.session_state.stats_computed:
        return st.session_state.analysis_results

    group_col = processed.structure.group_col
    if not group_col:
        return None

    report_gen = _ensure_report_generator(processed, settings)
    results = report_gen.run_all_statistics()
    st.session_state.analysis_results = results
    st.session_state.stats_computed = True
    st.session_state.stats_sections_done = {'individual_ba', 'totals', 'ratios', 'percentages', 'categories'}
    return results


def get_sig_pairs(result, max_pairs=5):
    """Extract significant pairs from result."""
    if not result or not result.posthoc_test or result.posthoc_test.pairwise_results is None:
        return []
    pairs = []
    df = result.posthoc_test.pairwise_results
    sig = df[df['significant']].sort_values('pvalue_adj' if 'pvalue_adj' in df else 'pvalue')
    for _, row in sig.head(max_pairs).iterrows():
        p = row.get('pvalue_adj', row.get('pvalue', 1.0))
        annot = '***' if p < 0.001 else ('**' if p < 0.01 else '*' if p < 0.05 else '')
        if annot:
            pairs.append((row['group1'], row['group2'], annot))
    return pairs


def generate_all_export_figures(processed, results, settings):
    """Generate all figure variations for export package."""
    from config.bile_acid_species import get_primary, get_secondary, get_glycine_conjugated, get_taurine_conjugated
    
    figures = {}
    viz = BileAcidVisualizer(color_palette=settings['color_palette'], style=settings['plot_style'])
    group_col = processed.structure.group_col
    
    if not group_col or not results:
        return figures
    
    mask = processed.sample_data[group_col].notna() & (processed.sample_data[group_col].astype(str).str.lower() != 'nan')
    data = processed.sample_data.loc[mask]
    available_bas = processed.structure.bile_acid_cols

    is_threeway = getattr(results, 'is_threeway', False)
    is_twoway = results.is_twoway

    if is_threeway:
        # =====================================================================
        # THREE-WAY ANOVA EXPORT FIGURES
        # =====================================================================
        fa_col, fb_col, fc_col = results.factor_a_col, results.factor_b_col, results.factor_c_col
        fa_name, fb_name, fc_name = results.factor_a_name, results.factor_b_name, results.factor_c_name

        conc_selections = {
            'top10': processed.concentrations.mean().nlargest(10).index.tolist(),
            'significant': [b for b in available_bas if b in results.threeway_individual_ba
                           and any(getattr(results.threeway_individual_ba[b].threeway_result, attr) < settings['alpha']
                                   for attr in ['factor_a_pvalue', 'factor_b_pvalue', 'factor_c_pvalue'])][:10],
            'primary': [b for b in get_primary() if b in available_bas][:10],
            'secondary': [b for b in get_secondary() if b in available_bas][:10],
        }

        for sel_name, selected in conc_selections.items():
            if not selected:
                continue
            tw_stats = {b: results.threeway_individual_ba.get(b) for b in selected
                       if b in results.threeway_individual_ba}
            suffix = f"_{sel_name}"
            try:
                fig = viz.plot_threeway_multi_panel(
                    data=data, value_cols=selected,
                    factor_a_col=fa_col, factor_b_col=fb_col, factor_c_col=fc_col,
                    threeway_results=tw_stats, ncols=3,
                    factor_a_name=fa_name, factor_b_name=fb_name, factor_c_name=fc_name,
                    ylabel='Concentration (nmol/L)', show_points=settings['show_points'],
                    plot_type=settings['plot_type'])
                figures[f'concentrations{suffix}'] = fig
            except Exception:
                pass

        key_totals = ['total_all', 'total_primary', 'total_secondary', 'total_conjugated',
                      'total_unconjugated', 'glycine_conjugated', 'taurine_conjugated',
                      'sulfated', 'oxidized', 'epimerized', 'nor_bile_acids',
                      'total_12alpha_hydroxylated', 'total_non12alpha_hydroxylated',
                      'primary_conjugated', 'secondary_conjugated',
                      'primary_unconjugated', 'secondary_unconjugated']
        available_totals = [t for t in key_totals if t in processed.totals.columns]

        if available_totals:
            totals_data = pd.concat([data[[fa_col, fb_col, fc_col]].reset_index(drop=True),
                                     processed.totals[available_totals].loc[data.index].reset_index(drop=True)], axis=1)
            tw_stats = {t: results.threeway_totals.get(t) for t in available_totals
                       if t in results.threeway_totals}
            try:
                fig = viz.plot_threeway_multi_panel(
                    data=totals_data, value_cols=available_totals,
                    factor_a_col=fa_col, factor_b_col=fb_col, factor_c_col=fc_col,
                    threeway_results=tw_stats, ncols=3,
                    factor_a_name=fa_name, factor_b_name=fb_name, factor_c_name=fc_name,
                    ylabel='Concentration (nmol/L)', show_points=settings['show_points'],
                    plot_type=settings['plot_type'])
                figures['totals'] = fig
            except Exception:
                pass

        ratio_cols = [col for col in processed.ratios.columns if not processed.ratios[col].isna().all()]
        if ratio_cols:
            ratio_data = pd.concat([data[[fa_col, fb_col, fc_col]].reset_index(drop=True),
                                    processed.ratios[ratio_cols].loc[data.index].reset_index(drop=True)], axis=1)
            tw_stats = {r: results.threeway_ratios.get(r) for r in ratio_cols[:9]
                       if r in results.threeway_ratios}
            try:
                fig = viz.plot_threeway_multi_panel(
                    data=ratio_data, value_cols=ratio_cols[:9],
                    factor_a_col=fa_col, factor_b_col=fb_col, factor_c_col=fc_col,
                    threeway_results=tw_stats, ncols=3,
                    factor_a_name=fa_name, factor_b_name=fb_name, factor_c_name=fc_name,
                    ylabel='Ratio', show_points=settings['show_points'],
                    plot_type=settings['plot_type'])
                figures['ratios'] = fig
            except Exception:
                pass

    elif is_twoway:
        # =====================================================================
        # TWO-WAY ANOVA EXPORT FIGURES
        # =====================================================================
        fa_col = results.factor_a_col
        fb_col = results.factor_b_col
        fa_name = results.factor_a_name
        fb_name = results.factor_b_name

        # --- Concentrations ---
        conc_selections = {
            'top10': processed.concentrations.mean().nlargest(10).index.tolist(),
            'significant': [b for b in available_bas if b in results.twoway_individual_ba
                           and (results.twoway_individual_ba[b].twoway_result.factor_a_pvalue < settings['alpha']
                                or results.twoway_individual_ba[b].twoway_result.factor_b_pvalue < settings['alpha']
                                or results.twoway_individual_ba[b].twoway_result.interaction_pvalue < settings['alpha'])][:10],
            'primary': [b for b in get_primary() if b in available_bas][:10],
            'secondary': [b for b in get_secondary() if b in available_bas][:10],
        }

        for sel_name, selected in conc_selections.items():
            if not selected:
                continue
            tw_stats = {b: results.twoway_individual_ba.get(b) for b in selected
                       if b in results.twoway_individual_ba}
            suffix = f"_{sel_name}"
            try:
                fig = viz.plot_twoway_multi_panel(
                    data=data, value_cols=selected,
                    factor_a_col=fa_col, factor_b_col=fb_col,
                    twoway_results=tw_stats, ncols=3,
                    factor_a_name=fa_name, factor_b_name=fb_name,
                    ylabel='Concentration (nmol/L)', show_points=settings['show_points'],
                    plot_type=settings['plot_type'])
                figures[f'concentrations{suffix}'] = fig
            except Exception:
                pass
            try:
                fig_int = viz.plot_twoway_interaction_multi_panel(
                    data=data, value_cols=selected,
                    factor_a_col=fa_col, factor_b_col=fb_col,
                    twoway_results=tw_stats, ncols=3,
                    factor_a_name=fa_name, factor_b_name=fb_name,
                    ylabel='Concentration (nmol/L)')
                figures[f'concentrations{suffix}_interaction'] = fig_int
            except Exception:
                pass

        # --- Totals ---
        key_totals = ['total_all', 'total_primary', 'total_secondary', 'total_conjugated',
                      'total_unconjugated', 'glycine_conjugated', 'taurine_conjugated',
                      'sulfated', 'oxidized', 'epimerized', 'nor_bile_acids',
                      'total_12alpha_hydroxylated', 'total_non12alpha_hydroxylated',
                      'primary_conjugated', 'secondary_conjugated',
                      'primary_unconjugated', 'secondary_unconjugated']
        available_totals = [t for t in key_totals if t in processed.totals.columns]

        if available_totals:
            totals_data = pd.concat([data[[fa_col, fb_col]].reset_index(drop=True),
                                     processed.totals[available_totals].loc[data.index].reset_index(drop=True)], axis=1)
            tw_stats = {t: results.twoway_totals.get(t) for t in available_totals
                       if t in results.twoway_totals}
            try:
                fig = viz.plot_twoway_multi_panel(
                    data=totals_data, value_cols=available_totals,
                    factor_a_col=fa_col, factor_b_col=fb_col,
                    twoway_results=tw_stats, ncols=3,
                    factor_a_name=fa_name, factor_b_name=fb_name,
                    ylabel='Concentration (nmol/L)', show_points=settings['show_points'],
                    plot_type=settings['plot_type'])
                figures['totals'] = fig
            except Exception:
                pass
            try:
                fig_int = viz.plot_twoway_interaction_multi_panel(
                    data=totals_data, value_cols=available_totals,
                    factor_a_col=fa_col, factor_b_col=fb_col,
                    twoway_results=tw_stats, ncols=3,
                    factor_a_name=fa_name, factor_b_name=fb_name,
                    ylabel='Concentration (nmol/L)')
                figures['totals_interaction'] = fig_int
            except Exception:
                pass

        # --- Ratios ---
        ratio_cols = [col for col in processed.ratios.columns if not processed.ratios[col].isna().all()]
        if ratio_cols:
            ratio_data = pd.concat([data[[fa_col, fb_col]].reset_index(drop=True),
                                    processed.ratios[ratio_cols].loc[data.index].reset_index(drop=True)], axis=1)
            tw_stats = {r: results.twoway_ratios.get(r) for r in ratio_cols[:9]
                       if r in results.twoway_ratios}
            try:
                fig = viz.plot_twoway_multi_panel(
                    data=ratio_data, value_cols=ratio_cols[:9],
                    factor_a_col=fa_col, factor_b_col=fb_col,
                    twoway_results=tw_stats, ncols=3,
                    factor_a_name=fa_name, factor_b_name=fb_name,
                    ylabel='Ratio', show_points=settings['show_points'],
                    plot_type=settings['plot_type'])
                figures['ratios'] = fig
            except Exception:
                pass

    else:
        # =====================================================================
        # ONE-WAY EXPORT FIGURES (unchanged)
        # =====================================================================

        # === CONCENTRATIONS TAB FIGURES ===
        conc_selections = {
            'top10': processed.concentrations.mean().nlargest(10).index.tolist(),
            'significant': [b for b in available_bas if b in results.individual_ba_results
                           and results.individual_ba_results[b].main_test.significant][:10],
            'primary': [b for b in get_primary() if b in available_bas][:10],
            'secondary': [b for b in get_secondary() if b in available_bas][:10],
        }

        for sel_name, selected in conc_selections.items():
            if not selected:
                continue
            stats_dict = {b: results.individual_ba_results.get(b) for b in selected if b in results.individual_ba_results}

            for log_scale in [False, True]:
                suffix = f"_{sel_name}{'_log' if log_scale else ''}"
                try:
                    fig = viz.plot_multi_panel_groups_with_stats(
                        data, selected, group_col, stats_dict, ncols=3,
                        plot_type=settings['plot_type'], log_scale=log_scale,
                        show_points=settings['show_points'], ylabel='Concentration (nmol/L)')
                    figures[f'concentrations{suffix}'] = fig
                except Exception:
                    pass

        # === TOTALS TAB FIGURES ===
        totals_combined = pd.concat([data[[group_col]], processed.totals.loc[data.index]], axis=1)

        key_totals = ['total_all', 'total_primary', 'total_secondary', 'total_conjugated',
                      'total_unconjugated', 'glycine_conjugated', 'taurine_conjugated',
                      'sulfated', 'oxidized', 'epimerized', 'nor_bile_acids',
                      'total_12alpha_hydroxylated', 'total_non12alpha_hydroxylated',
                      'primary_conjugated', 'secondary_conjugated',
                      'primary_unconjugated', 'secondary_unconjugated']
        available_totals = [t for t in key_totals if t in totals_combined.columns]

        if available_totals:
            totals_stats = {t: results.totals_results.get(t) for t in available_totals if t in results.totals_results}
            for log_scale in [False, True]:
                suffix = '_log' if log_scale else ''
                try:
                    fig = viz.plot_multi_panel_groups_with_stats(
                        totals_combined, available_totals, group_col, totals_stats, ncols=3,
                        plot_type=settings['plot_type'], log_scale=log_scale,
                        show_points=settings['show_points'], ylabel='Concentration (nmol/L)')
                    figures[f'totals{suffix}'] = fig
                except Exception:
                    pass

        # === PERCENTAGES TAB FIGURES ===
        pct_cols = [col for col in processed.percentages.columns if col.endswith('_pct')]
        if pct_cols:
            top_pct = processed.percentages[pct_cols].mean().nlargest(10).index.tolist()
            pct_combined = pd.concat([data[[group_col]], processed.percentages.loc[data.index, top_pct]], axis=1)

            display_cols = [col.replace('_pct', '') for col in top_pct]
            pct_display = pct_combined.rename(columns={old: new for old, new in zip(top_pct, display_cols)})

            pct_stats = {}
            for pct_col, disp_col in zip(top_pct, display_cols):
                if pct_col in results.percentages_results:
                    pct_stats[disp_col] = results.percentages_results[pct_col]

            try:
                fig = viz.plot_multi_panel_groups_with_stats(
                    pct_display, display_cols, group_col, pct_stats, ncols=3,
                    plot_type=settings['plot_type'], log_scale=False,
                    show_points=settings['show_points'], ylabel='% of Total BA')
                figures['percentages_top10'] = fig
            except Exception:
                pass

        # === RATIOS TAB FIGURES ===
        ratio_cols = [col for col in processed.ratios.columns if not processed.ratios[col].isna().all()]
        if ratio_cols:
            ratios_combined = pd.concat([data[[group_col]], processed.ratios.loc[data.index, ratio_cols]], axis=1)
            ratios_stats = {r: results.ratios_results.get(r) for r in ratio_cols if r in results.ratios_results}

            for log_scale in [False, True]:
                suffix = '_log' if log_scale else ''
                try:
                    fig = viz.plot_multi_panel_groups_with_stats(
                        ratios_combined, ratio_cols[:9], group_col, ratios_stats, ncols=3,
                        plot_type=settings['plot_type'], log_scale=log_scale,
                        show_points=settings['show_points'], ylabel='Ratio')
                    figures[f'ratios{suffix}'] = fig
                except Exception:
                    pass

    return figures


def create_results_zip(processed, results, figures, report_gen, settings=None):
    """Create ZIP with all results."""
    # Get metadata columns
    id_cols = []
    if processed.structure.sample_id_col and processed.structure.sample_id_col in processed.sample_data.columns:
        id_cols.append(processed.structure.sample_id_col)
    if processed.structure.group_col and processed.structure.group_col in processed.sample_data.columns:
        id_cols.append(processed.structure.group_col)
    # Include Factor_ columns
    for factor_col in processed.structure.factors.values():
        if factor_col in processed.sample_data.columns and factor_col not in id_cols:
            id_cols.append(factor_col)

    metadata = processed.sample_data[id_cols] if id_cols else pd.DataFrame(index=processed.sample_data.index)

    buf = BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        # Data files with metadata
        for name, df in [('concentrations', processed.concentrations),
                         ('percentages', processed.percentages),
                         ('totals', processed.totals),
                         ('ratios', processed.ratios)]:
            export_df = pd.concat([metadata, df], axis=1)
            zf.writestr(f'data/{name}.xlsx', df_to_excel_bytes(export_df))

        # Full data
        full = pd.concat([processed.sample_data, processed.totals, processed.ratios, processed.percentages], axis=1)
        zf.writestr('data/full_analysis.xlsx', df_to_excel_bytes(full))
        
        # LOD-highlighted Excel
        lod_handling = settings.get('lod_handling', 'half_lod') if settings else 'half_lod'
        lod_excel_bytes = create_full_data_excel_with_highlighting(processed, lod_handling)
        zf.writestr('data/full_data_lod_highlighted.xlsx', lod_excel_bytes)
        
        # Excel report
        if report_gen:
            excel_buf = BytesIO()
            report_gen.save_excel_report(excel_buf)
            excel_buf.seek(0)
            zf.writestr('reports/statistical_report.xlsx', excel_buf.getvalue())
        
        # Figures — may be dicts with png/pdf keys, raw bytes, or matplotlib objects
        for name, fig in figures.items():
            if fig:
                if isinstance(fig, dict):
                    if fig.get('png'):
                        zf.writestr(f'figures/{name}.png', fig['png'])
                    if fig.get('pdf'):
                        zf.writestr(f'figures/{name}.pdf', fig['pdf'])
                elif isinstance(fig, bytes):
                    zf.writestr(f'figures/{name}.png', fig)
                else:
                    zf.writestr(f'figures/{name}.png', fig_to_bytes(fig, 'png'))
                    zf.writestr(f'figures/{name}.pdf', fig_to_bytes(fig, 'pdf'))
    
    buf.seek(0)
    return buf.getvalue()


def create_full_data_excel_with_highlighting(processed, lod_handling='half_lod') -> bytes:
    """Create Excel workbook with yellow highlighting on cells where LOD values were replaced."""
    from openpyxl.styles import PatternFill
    
    lod_display = { "lod": "LOD value","half_lod": "LOD/2", "zero": "Zero", "drop": "NaN (excluded)"}
    
    buf = BytesIO()
    
    # Get metadata columns
    id_cols = []
    if processed.structure.sample_id_col and processed.structure.sample_id_col in processed.sample_data.columns:
        id_cols.append(processed.structure.sample_id_col)
    if processed.structure.group_col and processed.structure.group_col in processed.sample_data.columns:
        id_cols.append(processed.structure.group_col)
    # Include Factor_ columns
    for factor_col in processed.structure.factors.values():
        if factor_col in processed.sample_data.columns and factor_col not in id_cols:
            id_cols.append(factor_col)

    metadata = processed.sample_data[id_cols] if id_cols else pd.DataFrame(index=processed.sample_data.index)

    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        # Concentrations sheet with highlighting
        conc_df = pd.concat([metadata, processed.concentrations], axis=1)
        conc_df.to_excel(writer, sheet_name='Concentrations', index=False)
        
        # Totals
        totals_df = pd.concat([metadata, processed.totals], axis=1)
        totals_df.to_excel(writer, sheet_name='Totals', index=False)
        
        # Ratios
        ratios_df = pd.concat([metadata, processed.ratios], axis=1)
        ratios_df.to_excel(writer, sheet_name='Ratios', index=False)
        
        # Percentages
        pct_df = pd.concat([metadata, processed.percentages], axis=1)
        pct_df.to_excel(writer, sheet_name='Percentages', index=False)
        
        # LOD Summary sheet
        lod_summary_rows = []
        for ba in processed.structure.bile_acid_cols:
            lod_val = processed.structure.analyte_lods.get(ba, 'N/A')
            count = processed.structure.analyte_lod_counts.get(ba, 0)
            n = len(processed.sample_data)
            
            # Compute replacement value based on handling method
            if isinstance(lod_val, (int, float)):
                if lod_handling == 'half_lod':
                    repl = lod_val / 2
                elif lod_handling == 'lod':
                    repl = lod_val
                elif lod_handling == 'zero':
                    repl = 0
                else:
                    repl = 'NaN'
            else:
                repl = 'N/A'
            
            lod_summary_rows.append({
                'Analyte': ba,
                'LOD (nM)': lod_val,
                'Replacement Value': repl,
                'Samples Below LOD': count,
                '% Below LOD': f"{count/n*100:.1f}" if n > 0 else "0",
                'Total Samples': n
            })
        lod_df = pd.DataFrame(lod_summary_rows)
        lod_df.to_excel(writer, sheet_name='LOD Summary', index=False)
        
        # Legend sheet
        legend_data = pd.DataFrame({
            'Item': ['Yellow cells', 'LOD Source', 'LOD Handling'],
            'Description': [
                f'Value was below LOD and replaced with {lod_display.get(lod_handling, lod_handling)}',
                processed.structure.lod_source,
                f'{lod_handling} ({lod_display.get(lod_handling, lod_handling)})'
            ]
        })
        legend_data.to_excel(writer, sheet_name='Legend', index=False)
        
        # Apply yellow highlighting to LOD-replaced cells on Concentrations sheet
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        ws = writer.sheets['Concentrations']
        
        # Map column names to Excel column indices (1-based, after metadata columns)
        col_headers = list(conc_df.columns)
        for ba, row_indices in processed.structure.analyte_lod_rows.items():
            if ba in col_headers:
                col_idx = col_headers.index(ba) + 1  # 1-based for openpyxl
                for row_idx in row_indices:
                    # +2 because Excel is 1-based and row 1 is the header
                    ws.cell(row=row_idx + 2, column=col_idx).fill = yellow_fill
    
    buf.seek(0)
    return buf.getvalue()


def render_sidebar():
    """Render sidebar settings."""
    st.sidebar.markdown("## ⚙️ Settings")
    
    st.sidebar.markdown("### 📊 Detection Limits")
    lod_handling = st.sidebar.selectbox(
        "Below-LOD handling",
        [ "lod", "half_lod", "zero", "drop"],
        format_func=lambda x: { "lod": "LOD value (recommended)", "half_lod": "LOD/2",
                               "zero": "Zero", "drop": "NaN (exclude)"}[x],
        help="LOD values are auto-detected per analyte from the LC-MS standards sheet"
    )
    lod_threshold = st.sidebar.slider(
        "LOD exclusion threshold (%)",
        min_value=0, max_value=100, value=50, step=5,
        help="Exclude analytes from statistical testing if ≥ this % of values "
             "are below LOD. Set to 0 to disable. Uses modified rule: analyte is "
             "kept if ANY group has ≥ (100 − threshold)% detected values."
    )
    
    st.sidebar.markdown("### 📈 Analysis")
    alpha = st.sidebar.slider("Significance (α)", 0.01, 0.10, 0.05, 0.01)
    plot_type = st.sidebar.selectbox("Plot type", ["box", "violin", "bar", "strip"])
    show_points = st.sidebar.checkbox("Show data points", True)
    
    # Color palette selection
    st.sidebar.markdown("### 🎨 Appearance")
    palette_options = {
        "Set2": "Set2 (Default - Soft pastels)",
        "Set1": "Set1 (Bold primary)",
        "Paired": "Paired (Light/dark pairs)",
        "Dark2": "Dark2 (Darker pastels)",
        "colorblind": "Colorblind-friendly",
        "tab10": "Tab10 (Matplotlib default)",
        "Pastel1": "Pastel1 (Very soft)",
        "Accent": "Accent (High contrast)"
    }
    color_palette = st.sidebar.selectbox("Color palette", list(palette_options.keys()),
                                         format_func=lambda x: palette_options[x])
    
    # Background style selection
    style_options = {
        "whitegrid": "White with grid",
        "white": "Clean white",
        "darkgrid": "Gray with grid",
        "ticks": "White with ticks"
    }
    plot_style = st.sidebar.selectbox("Plot background", list(style_options.keys()),
                                      format_func=lambda x: style_options[x])
    
    return {'lod_handling': lod_handling, 'lod_threshold': lod_threshold,
            'alpha': alpha, 'plot_type': plot_type,
            'show_points': show_points,
            'color_palette': color_palette, 'plot_style': plot_style}


def render_concentrations_tab(processed, settings):
    """Render concentrations tab."""
    st.markdown("### Individual Bile Acid Concentrations")

    viz = BileAcidVisualizer(color_palette=settings['color_palette'], style=settings['plot_style'])
    group_col = processed.structure.group_col
    results = st.session_state.analysis_results

    if not group_col:
        st.warning("No group column detected.")
        return

    mask = processed.sample_data[group_col].notna() & (processed.sample_data[group_col].astype(str).str.lower() != 'nan')
    data = processed.sample_data.loc[mask]
    available_bas = processed.structure.bile_acid_cols
    is_threeway = results is not None and getattr(results, 'is_threeway', False)
    is_twoway = results is not None and results.is_twoway

    quick = st.segmented_control("Quick select",
        ["Top 10", "Significant", "Primary", "Secondary", "Conjugated", "Unconjugated",
         "Glycine", "Taurine", "Sulfated", "Oxidized", "Epimerized",
         "12α-OH", "Non-12α-OH", "Nor", "Custom"],
        default="Top 10")

    if quick == "Top 10":
        selected = processed.concentrations.mean().nlargest(10).index.tolist()
    elif quick == "Significant":
        if is_threeway:
            selected = [b for b in available_bas if b in results.threeway_individual_ba
                       and any(getattr(results.threeway_individual_ba[b].threeway_result, attr) < settings['alpha']
                               for attr in ['factor_a_pvalue', 'factor_b_pvalue', 'factor_c_pvalue'])][:10]
        elif is_twoway:
            selected = [b for b in available_bas if b in results.twoway_individual_ba
                       and (results.twoway_individual_ba[b].twoway_result.factor_a_pvalue < settings['alpha']
                            or results.twoway_individual_ba[b].twoway_result.factor_b_pvalue < settings['alpha']
                            or results.twoway_individual_ba[b].twoway_result.interaction_pvalue < settings['alpha'])][:10]
        else:
            selected = [b for b in available_bas if b in results.individual_ba_results
                       and results.individual_ba_results[b].main_test.significant][:10]
        if not selected:
            st.info("No significant individual BAs found.")
            selected = processed.concentrations.mean().nlargest(5).index.tolist()
    elif quick == "Primary":
        selected = [b for b in get_primary() if b in available_bas][:10]
    elif quick == "Secondary":
        selected = [b for b in get_secondary() if b in available_bas][:10]
    elif quick == "Conjugated":
        selected = [b for b in get_conjugated() if b in available_bas][:10]
    elif quick == "Unconjugated":
        selected = [b for b in get_unconjugated() if b in available_bas][:10]
    elif quick == "Glycine":
        selected = [b for b in get_glycine_conjugated() if b in available_bas][:10]
    elif quick == "Taurine":
        selected = [b for b in get_taurine_conjugated() if b in available_bas][:10]
    elif quick == "Sulfated":
        selected = [b for b in get_sulfated() if b in available_bas][:10]
    elif quick == "Oxidized":
        selected = [b for b in get_keto_derivatives() if b in available_bas][:10]
    elif quick == "Epimerized":
        selected = [b for b in get_iso_forms() if b in available_bas][:10]
    elif quick == "12α-OH":
        selected = [b for b in get_12alpha_hydroxylated() if b in available_bas][:10]
    elif quick == "Non-12α-OH":
        selected = [b for b in get_non12alpha_hydroxylated() if b in available_bas][:10]
    elif quick == "Nor":
        selected = [b for b in get_nor_bile_acids() if b in available_bas][:10]
    else:
        selected = st.multiselect("Select BAs", available_bas, available_bas[:5])

    log_scale = st.checkbox("Log10 scale", key="conc_log")

    # Show LOD exclusion notice if any selected analytes were excluded
    lod_excluded = getattr(results, 'lod_excluded', {})
    if lod_excluded and selected:
        excluded_selected = {ba: pct for ba, pct in lod_excluded.items() if ba in selected}
        if excluded_selected:
            excl_list = ", ".join(f"{ba} ({pct}%)" for ba, pct in sorted(excluded_selected.items()))
            st.info(
                f"**LOD exclusion** (threshold: ≥{getattr(results, 'lod_threshold', 50)}% below LOD): "
                f"The following analytes were excluded from statistical testing: {excl_list}. "
                f"Plots are still shown but without ANOVA results."
            )

    if selected:
        if is_threeway:
            # === THREE-WAY ANOVA PLOTS ===
            fa_col, fb_col, fc_col = results.factor_a_col, results.factor_b_col, results.factor_c_col
            fa_name, fb_name, fc_name = results.factor_a_name, results.factor_b_name, results.factor_c_name

            st.markdown(f"**Three-way ANOVA**: {fa_name} x {fb_name} x {fc_name}")

            tw_stats = {b: results.threeway_individual_ba.get(b) for b in selected
                       if b in results.threeway_individual_ba}

            fig = viz.plot_threeway_multi_panel(
                data=data, value_cols=selected,
                factor_a_col=fa_col, factor_b_col=fb_col, factor_c_col=fc_col,
                threeway_results=tw_stats, ncols=3,
                factor_a_name=fa_name, factor_b_name=fb_name, factor_c_name=fc_name,
                ylabel='Concentration (nmol/L)', show_points=settings['show_points'],
                plot_type=settings['plot_type']
            )
            st.pyplot(fig)
            store_figure(fig, 'concentrations_threeway')
            plt.close(fig)

            with st.expander("Interaction Plots"):
                fig_int = viz.plot_threeway_interaction_multi_panel(
                    data=data, value_cols=selected,
                    factor_a_col=fa_col, factor_b_col=fb_col, factor_c_col=fc_col,
                    threeway_results=tw_stats, ncols=3,
                    factor_a_name=fa_name, factor_b_name=fb_name, factor_c_name=fc_name,
                    ylabel='Concentration (nmol/L)'
                )
                st.pyplot(fig_int)
                store_figure(fig_int, 'concentrations_threeway_interaction')
                plt.close(fig_int)

            with st.expander("Three-Way ANOVA Summary"):
                if tw_stats:
                    st.dataframe(get_threeway_differences_summary(tw_stats), hide_index=True)
        elif is_twoway:
            # === TWO-WAY ANOVA PLOTS ===
            fa_col = results.factor_a_col
            fb_col = results.factor_b_col
            fa_name = results.factor_a_name
            fb_name = results.factor_b_name

            st.markdown(f"**Two-way ANOVA**: {fa_name} x {fb_name}")

            tw_stats = {b: results.twoway_individual_ba.get(b) for b in selected
                       if b in results.twoway_individual_ba}

            fig = viz.plot_twoway_multi_panel(
                data=data, value_cols=selected,
                factor_a_col=fa_col, factor_b_col=fb_col,
                twoway_results=tw_stats, ncols=3,
                factor_a_name=fa_name, factor_b_name=fb_name,
                ylabel='Concentration (nmol/L)', show_points=settings['show_points'],
                plot_type=settings['plot_type']
            )
            st.pyplot(fig)
            store_figure(fig, 'concentrations_twoway')
            plt.close(fig)

            with st.expander("Interaction Plots"):
                fig_int = viz.plot_twoway_interaction_multi_panel(
                    data=data, value_cols=selected,
                    factor_a_col=fa_col, factor_b_col=fb_col,
                    twoway_results=tw_stats, ncols=3,
                    factor_a_name=fa_name, factor_b_name=fb_name,
                    ylabel='Concentration (nmol/L)'
                )
                st.pyplot(fig_int)
                store_figure(fig_int, 'concentrations_interaction')
                plt.close(fig_int)

            with st.expander("Two-Way ANOVA Summary"):
                if tw_stats:
                    st.dataframe(get_twoway_differences_summary(tw_stats), hide_index=True)
        else:
            # === ONE-WAY ANOVA PLOTS ===
            stats_dict = {b: results.individual_ba_results.get(b) for b in selected if b in results.individual_ba_results}

            fig = viz.plot_multi_panel_groups_with_stats(data, selected, group_col, stats_dict,
                                                         ncols=3, plot_type=settings['plot_type'], log_scale=log_scale,
                                                         show_points=settings['show_points'], ylabel='Concentration (nmol/L)')
            st.pyplot(fig)
            store_figure(fig, f'concentrations{"_log" if log_scale else ""}')
            plt.close(fig)

            fig_other = viz.plot_multi_panel_groups_with_stats(data, selected, group_col, stats_dict,
                                                               ncols=3, plot_type=settings['plot_type'], log_scale=not log_scale,
                                                               show_points=settings['show_points'], ylabel='Concentration (nmol/L)')
            store_figure(fig_other, f'concentrations{"_log" if not log_scale else ""}')
            plt.close(fig_other)

            with st.expander("Statistical Summary"):
                st.dataframe(get_significant_differences_summary(stats_dict), hide_index=True)


def render_totals_tab(processed, settings):
    """Render totals tab."""
    st.markdown("### Aggregate Totals")

    viz = BileAcidVisualizer(color_palette=settings['color_palette'], style=settings['plot_style'])
    group_col = processed.structure.group_col
    results = st.session_state.analysis_results

    if not group_col:
        return

    # Only keep columns we actually need (factor/group cols) — avoid copying all BA columns
    mask = processed.sample_data[group_col].notna() & (processed.sample_data[group_col].astype(str).str.lower() != 'nan')
    data = processed.sample_data.loc[mask]
    is_threeway = results is not None and getattr(results, 'is_threeway', False)
    is_twoway = results is not None and results.is_twoway

    log_scale = st.checkbox("Log10 scale", key="totals_log")

    key_totals = ['total_all', 'total_primary', 'total_secondary', 'total_conjugated',
                  'total_unconjugated', 'glycine_conjugated', 'taurine_conjugated',
                  'sulfated', 'oxidized', 'epimerized', 'nor_bile_acids',
                  'total_12alpha_hydroxylated', 'total_non12alpha_hydroxylated',
                  'primary_conjugated', 'secondary_conjugated',
                  'primary_unconjugated', 'secondary_unconjugated']

    if is_threeway:
        fa_col, fb_col, fc_col = results.factor_a_col, results.factor_b_col, results.factor_c_col
        fa_name, fb_name, fc_name = results.factor_a_name, results.factor_b_name, results.factor_c_name
        combined = pd.concat([data[[fa_col, fb_col, fc_col]], processed.totals.loc[data.index]], axis=1)
        available = [t for t in key_totals if t in combined.columns]

        tw_stats = {t: results.threeway_totals.get(t) for t in available if t in results.threeway_totals}

        fig = viz.plot_threeway_multi_panel(
            data=combined, value_cols=available,
            factor_a_col=fa_col, factor_b_col=fb_col, factor_c_col=fc_col,
            threeway_results=tw_stats, ncols=3,
            factor_a_name=fa_name, factor_b_name=fb_name, factor_c_name=fc_name,
            ylabel='Concentration (nmol/L)', show_points=settings['show_points'],
            plot_type=settings['plot_type']
        )
        st.pyplot(fig)
        store_figure(fig, 'totals')
        plt.close(fig)
        del fig, combined
        gc.collect()

        with st.expander("Three-Way ANOVA Summary"):
            if tw_stats:
                st.dataframe(get_threeway_differences_summary(tw_stats), hide_index=True)
    elif is_twoway:
        fa_col = results.factor_a_col
        fb_col = results.factor_b_col
        combined = pd.concat([data[[fa_col, fb_col]], processed.totals.loc[data.index]], axis=1)
        available = [t for t in key_totals if t in combined.columns]

        tw_stats = {t: results.twoway_totals.get(t) for t in available if t in results.twoway_totals}

        fig = viz.plot_twoway_multi_panel(
            data=combined, value_cols=available,
            factor_a_col=fa_col, factor_b_col=fb_col,
            twoway_results=tw_stats, ncols=3,
            factor_a_name=results.factor_a_name, factor_b_name=results.factor_b_name,
            ylabel='Concentration (nmol/L)', show_points=settings['show_points'],
            plot_type=settings['plot_type']
        )
        st.pyplot(fig)
        store_figure(fig, 'totals')
        plt.close(fig)

        with st.expander("Two-Way ANOVA Summary"):
            if tw_stats:
                st.dataframe(get_twoway_differences_summary(tw_stats), hide_index=True)
    else:
        combined = pd.concat([data[[group_col]], processed.totals.loc[data.index]], axis=1)
        available = [t for t in key_totals if t in combined.columns]

        stats_dict = {t: results.totals_results.get(t) for t in available if t in results.totals_results}

        fig = viz.plot_multi_panel_groups_with_stats(combined, available, group_col, stats_dict,
                                                     ncols=3, plot_type=settings['plot_type'], log_scale=log_scale,
                                                     show_points=settings['show_points'], ylabel='Concentration (nmol/L)')
        st.pyplot(fig)
        store_figure(fig, f'totals{"_log" if log_scale else ""}')
        plt.close(fig)

        fig_other = viz.plot_multi_panel_groups_with_stats(combined, available, group_col, stats_dict,
                                                           ncols=3, plot_type=settings['plot_type'], log_scale=not log_scale,
                                                           show_points=settings['show_points'], ylabel='Concentration (nmol/L)')
        store_figure(fig_other, f'totals{"_log" if not log_scale else ""}')
        plt.close(fig_other)

        with st.expander("Statistical Summary"):
            rows = []
            for t in available:
                res = results.totals_results.get(t)
                if res:
                    rows.append({
                        'Total': t.replace('_', ' ').title(),
                        'P-value': f"{res.main_test.pvalue:.4f}",
                        'Significant': '\u2713' if res.main_test.significant else '',
                        'APA': format_apa_statistics(res)
                    })
            st.dataframe(pd.DataFrame(rows), hide_index=True)

    # =========================================================================
    # Category Composition Pie Charts (rendered on demand to save memory)
    # =========================================================================
    st.markdown("---")
    st.markdown("#### Category Composition - Pie Charts")
    if st.button("🥧 Generate Pie Charts", key="gen_cat_pie",
                 help="Within-category breakdown: individual bile acids as % of each category total"):
        ba_names = processed.structure.bile_acid_cols
        if is_threeway:
            cat_pie_data = data.assign(
                _factorial_group_=data[results.factor_a_col].astype(str) + ' - '
                + data[results.factor_b_col].astype(str) + ' - '
                + data[results.factor_c_col].astype(str)
            )
            cat_group_col = '_factorial_group_'
        elif is_twoway:
            cat_pie_data = data.assign(
                _factorial_group_=data[results.factor_a_col].astype(str) + ' - '
                + data[results.factor_b_col].astype(str)
            )
            cat_group_col = '_factorial_group_'
        else:
            cat_pie_data = data
            cat_group_col = group_col

        fig_cat_pie = viz.plot_category_composition_pie_charts(
            data=cat_pie_data,
            group_col=cat_group_col,
            bile_acid_cols=ba_names,
            title='Bile Acid Category Composition by Group',
        )
        st.pyplot(fig_cat_pie)
        store_figure(fig_cat_pie, 'category_composition_pie')
        plt.close(fig_cat_pie)


def render_percentages_tab(processed, settings):
    """Render percentages tab - comparing % composition across groups."""
    st.markdown("### Bile Acid Pool Composition (%)")
    st.caption("Compare the percentage each bile acid contributes to the total pool across groups")
    
    viz = BileAcidVisualizer(color_palette=settings['color_palette'], style=settings['plot_style'])
    group_col = processed.structure.group_col
    results = st.session_state.analysis_results
    
    if not group_col:
        st.warning("No group column detected.")
        return
    
    mask = processed.sample_data[group_col].notna() & (processed.sample_data[group_col].astype(str).str.lower() != 'nan')
    data = processed.sample_data.loc[mask]
    percentages = processed.percentages.loc[data.index].copy()
    
    # Get available percentage columns (remove _pct suffix for display)
    pct_cols = [c for c in percentages.columns if c.endswith('_pct')]
    ba_names = [c.replace('_pct', '') for c in pct_cols]
    is_threeway = results is not None and getattr(results, 'is_threeway', False)
    is_twoway = results is not None and results.is_twoway

    quick = st.pills("Quick select",
                     ["Top 10 by mean %", "Significant", "Primary", "Secondary",
                      "Conjugated", "Unconjugated", "Glycine", "Taurine",
                      "Sulfated", "Oxidized", "Epimerized",
                      "12α-OH", "Non-12α-OH", "Nor", "Custom"],
                     default="Top 10 by mean %", key="pct_quick")

    def _pct_select(getter):
        return [f'{ba}_pct' for ba in getter() if f'{ba}_pct' in pct_cols][:10]

    if quick == "Top 10 by mean %":
        top_pcts = percentages[pct_cols].mean().nlargest(10).index.tolist()
        selected_pct_cols = top_pcts
    elif quick == "Significant":
        if is_threeway:
            selected_pct_cols = [c for c in pct_cols if c in results.threeway_percentages
                               and any(getattr(results.threeway_percentages[c].threeway_result, attr) < settings['alpha']
                                       for attr in ['factor_a_pvalue', 'factor_b_pvalue', 'factor_c_pvalue'])][:10]
        elif is_twoway:
            selected_pct_cols = [c for c in pct_cols if c in results.twoway_percentages
                               and (results.twoway_percentages[c].twoway_result.factor_a_pvalue < settings['alpha']
                                    or results.twoway_percentages[c].twoway_result.factor_b_pvalue < settings['alpha']
                                    or results.twoway_percentages[c].twoway_result.interaction_pvalue < settings['alpha'])][:10]
        else:
            selected_pct_cols = [c for c in pct_cols if c in results.percentages_results
                               and results.percentages_results[c].main_test.significant][:10]
        if not selected_pct_cols:
            st.info("No significant percentage differences found. Showing top 10 by mean %.")
            selected_pct_cols = percentages[pct_cols].mean().nlargest(10).index.tolist()
    elif quick == "Primary":
        selected_pct_cols = _pct_select(get_primary)
    elif quick == "Secondary":
        selected_pct_cols = _pct_select(get_secondary)
    elif quick == "Conjugated":
        selected_pct_cols = _pct_select(get_conjugated)
    elif quick == "Unconjugated":
        selected_pct_cols = _pct_select(get_unconjugated)
    elif quick == "Glycine":
        selected_pct_cols = _pct_select(get_glycine_conjugated)
    elif quick == "Taurine":
        selected_pct_cols = _pct_select(get_taurine_conjugated)
    elif quick == "Sulfated":
        selected_pct_cols = _pct_select(get_sulfated)
    elif quick == "Oxidized":
        selected_pct_cols = _pct_select(get_keto_derivatives)
    elif quick == "Epimerized":
        selected_pct_cols = _pct_select(get_iso_forms)
    elif quick == "12α-OH":
        selected_pct_cols = _pct_select(get_12alpha_hydroxylated)
    elif quick == "Non-12α-OH":
        selected_pct_cols = _pct_select(get_non12alpha_hydroxylated)
    elif quick == "Nor":
        selected_pct_cols = _pct_select(get_nor_bile_acids)
    else:
        selected_bas = st.multiselect("Select bile acids", ba_names, ba_names[:5], key="pct_custom")
        selected_pct_cols = [f'{ba}_pct' for ba in selected_bas]
    
    if not selected_pct_cols:
        st.warning("No bile acids selected.")
        return

    # Show significant findings summary
    if is_threeway:
        sig_pcts = [c for c in selected_pct_cols if c in results.threeway_percentages
                   and any(getattr(results.threeway_percentages[c].threeway_result, attr) < settings['alpha']
                           for attr in ['factor_a_pvalue', 'factor_b_pvalue', 'factor_c_pvalue'])]
    elif is_twoway:
        sig_pcts = [c for c in selected_pct_cols if c in results.twoway_percentages
                   and (results.twoway_percentages[c].twoway_result.factor_a_pvalue < settings['alpha']
                        or results.twoway_percentages[c].twoway_result.factor_b_pvalue < settings['alpha']
                        or results.twoway_percentages[c].twoway_result.interaction_pvalue < settings['alpha'])]
    else:
        sig_pcts = [c for c in selected_pct_cols if c in results.percentages_results
                   and results.percentages_results[c].main_test.significant]
    if sig_pcts:
        sig_names = [c.replace('_pct', '') for c in sig_pcts]
        st.success(f"**Significant differences:** {', '.join(sig_names)}")

    # Rename columns for cleaner display (remove _pct suffix)
    display_cols = {col: col.replace('_pct', '') for col in selected_pct_cols}
    display_names = list(display_cols.values())

    # =========================================================================
    # SECTION 1: Statistical comparison (box/violin plots)
    # =========================================================================
    st.markdown("#### Statistical Comparison by Group")

    if is_threeway:
        fa_col, fb_col, fc_col = results.factor_a_col, results.factor_b_col, results.factor_c_col
        fa_name, fb_name, fc_name = results.factor_a_name, results.factor_b_name, results.factor_c_name

        st.markdown(f"**Three-way ANOVA**: {fa_name} x {fb_name} x {fc_name}")

        tw_stats = {c.replace('_pct', ''): results.threeway_percentages.get(c)
                   for c in selected_pct_cols if c in results.threeway_percentages}

        pct_plot_data = pd.concat([data[[fa_col, fb_col, fc_col]].reset_index(drop=True),
                                   percentages[selected_pct_cols].reset_index(drop=True)], axis=1)
        pct_plot_data = pct_plot_data.rename(columns=display_cols)

        try:
            fig = viz.plot_threeway_multi_panel(
                data=pct_plot_data, value_cols=display_names,
                factor_a_col=fa_col, factor_b_col=fb_col, factor_c_col=fc_col,
                threeway_results=tw_stats, ncols=3,
                factor_a_name=fa_name, factor_b_name=fb_name, factor_c_name=fc_name,
                ylabel='% of Total BA', show_points=settings['show_points'],
                plot_type=settings['plot_type']
            )
            st.pyplot(fig)
            store_figure(fig, 'percentages_threeway')
            plt.close(fig)
        except Exception as e:
            st.warning(f"Could not generate three-way percentage plot: {e}")

        with st.expander("Three-Way ANOVA Summary"):
            tw_stats_orig = {c: results.threeway_percentages.get(c)
                           for c in selected_pct_cols if c in results.threeway_percentages}
            if tw_stats_orig:
                st.dataframe(get_threeway_differences_summary(tw_stats_orig), hide_index=True)
    elif is_twoway:
        fa_col = results.factor_a_col
        fb_col = results.factor_b_col
        fa_name = results.factor_a_name
        fb_name = results.factor_b_name

        st.markdown(f"**Two-way ANOVA**: {fa_name} x {fb_name}")

        tw_stats = {c.replace('_pct', ''): results.twoway_percentages.get(c)
                   for c in selected_pct_cols if c in results.twoway_percentages}

        pct_plot_data = pd.concat([data[[fa_col, fb_col]].reset_index(drop=True),
                                   percentages[selected_pct_cols].reset_index(drop=True)], axis=1)
        pct_plot_data = pct_plot_data.rename(columns=display_cols)

        try:
            fig = viz.plot_twoway_multi_panel(
                data=pct_plot_data, value_cols=display_names,
                factor_a_col=fa_col, factor_b_col=fb_col,
                twoway_results=tw_stats, ncols=3,
                factor_a_name=fa_name, factor_b_name=fb_name,
                ylabel='% of Total BA', show_points=settings['show_points'],
                plot_type=settings['plot_type']
            )
            st.pyplot(fig)
            store_figure(fig, 'percentages_twoway')
            plt.close(fig)
        except Exception as e:
            st.warning(f"Could not generate two-way percentage plot: {e}")

        with st.expander("Two-Way ANOVA Summary"):
            tw_stats_orig = {c: results.twoway_percentages.get(c)
                           for c in selected_pct_cols if c in results.twoway_percentages}
            if tw_stats_orig:
                st.dataframe(get_twoway_differences_summary(tw_stats_orig), hide_index=True)
    else:
        # Combine data for plotting
        plot_df = pd.concat([data[[group_col]].reset_index(drop=True),
                             percentages[selected_pct_cols].reset_index(drop=True)], axis=1)
        plot_df_display = plot_df.rename(columns=display_cols)

        # Get stats results with display names
        stats_dict = {}
        for pct_col in selected_pct_cols:
            display_name = pct_col.replace('_pct', '')
            if pct_col in results.percentages_results:
                stats_dict[display_name] = results.percentages_results[pct_col]

        fig = viz.plot_multi_panel_groups_with_stats(
            plot_df_display, display_names, group_col, stats_dict,
            ncols=3, plot_type=settings['plot_type'], log_scale=False,
            show_points=settings['show_points'], ylabel='% of Total BA'
        )

        st.pyplot(fig)
        store_figure(fig, 'percentages')
        plt.close(fig)

    # =========================================================================
    # SECTION 2: Pie Charts - Pool Composition
    # =========================================================================
    st.markdown("---")
    st.markdown("#### Pool Composition - Pie Charts")
    st.caption("Visual breakdown of bile acid pool for each group (top contributors)")

    if is_threeway:
        # Create combined group column for composition charts
        comp_group_col = '_factorial_group_'
        comp_data = data.copy()
        comp_data[comp_group_col] = (comp_data[results.factor_a_col].astype(str) + ' - '
                                     + comp_data[results.factor_b_col].astype(str) + ' - '
                                     + comp_data[results.factor_c_col].astype(str))
        pie_df = pd.concat([comp_data[[comp_group_col]].reset_index(drop=True),
                           percentages[pct_cols].reset_index(drop=True)], axis=1)
    elif is_twoway:
        # Create combined group column for composition charts
        comp_group_col = '_factorial_group_'
        comp_data = data.copy()
        comp_data[comp_group_col] = comp_data[results.factor_a_col].astype(str) + ' - ' + comp_data[results.factor_b_col].astype(str)
        pie_df = pd.concat([comp_data[[comp_group_col]].reset_index(drop=True),
                           percentages[pct_cols].reset_index(drop=True)], axis=1)
    else:
        comp_group_col = group_col
        pie_df = pd.concat([data[[group_col]].reset_index(drop=True),
                           percentages[pct_cols].reset_index(drop=True)], axis=1)

    fig_pie = viz.plot_composition_pie_charts(
        pie_df, comp_group_col, pct_cols,
        title='Bile Acid Pool Composition by Group',
        top_n=7, other_threshold=2.0
    )
    st.pyplot(fig_pie)
    store_figure(fig_pie, 'percentages_pie')
    plt.close(fig_pie)

    # =========================================================================
    # SECTION 3: Horizontal Bar Charts - Group Comparison
    # =========================================================================
    st.markdown("---")
    st.markdown("#### Pool Composition - Bar Comparison")
    st.caption("Side-by-side comparison of bile acid percentages across groups")
    
    fig_hbar = viz.plot_composition_horizontal_bars(
        pie_df, comp_group_col, pct_cols,
        title='Bile Acid Pool Composition Comparison',
        top_n=15, show_values=True
    )
    st.pyplot(fig_hbar)
    store_figure(fig_hbar, 'percentages_bars')
    plt.close(fig_hbar)
    
    # =========================================================================
    # SECTION 4: Stacked Bar - Full Composition
    # =========================================================================
    st.markdown("---")
    st.markdown("#### Stacked Composition View")
    st.caption("Full bile acid pool breakdown for each group")
    
    fig_stacked = viz.plot_composition_stacked_horizontal(
        pie_df, comp_group_col, pct_cols,
        title='Complete Bile Acid Pool Composition',
        top_n=12
    )
    st.pyplot(fig_stacked)
    store_figure(fig_stacked, 'percentages_stacked')
    plt.close(fig_stacked)
    
    # =========================================================================
    # SECTION 5: Statistical Summary Table
    # =========================================================================
    with st.expander("📊 Statistical Summary"):
        if is_threeway:
            tw_stats_all = {c: results.threeway_percentages.get(c)
                           for c in selected_pct_cols if c in results.threeway_percentages}
            if tw_stats_all:
                st.dataframe(get_threeway_differences_summary(tw_stats_all), hide_index=True)
        elif is_twoway:
            tw_stats_all = {c: results.twoway_percentages.get(c)
                           for c in selected_pct_cols if c in results.twoway_percentages}
            if tw_stats_all:
                st.dataframe(get_twoway_differences_summary(tw_stats_all), hide_index=True)
        else:
            rows = []
            for pct_col in selected_pct_cols:
                ba_name = pct_col.replace('_pct', '')
                res = results.percentages_results.get(pct_col)
                if res:
                    n_sig = 0
                    if res.posthoc_test and res.posthoc_test.pairwise_results is not None:
                        n_sig = res.posthoc_test.pairwise_results['significant'].sum()
                    rows.append({
                        'Bile Acid': ba_name,
                        'Test': res.main_test.test_type.value,
                        'P-value': f"{res.main_test.pvalue:.4f}",
                        'Significant': '✓' if res.main_test.significant else '',
                        'Effect Size': f"{res.main_test.effect_size:.3f}" if res.main_test.effect_size else 'N/A',
                        'Sig. Pairs': n_sig,
                        'APA': format_apa_statistics(res)
                    })
            if rows:
                st.dataframe(pd.DataFrame(rows), hide_index=True)

    # Group mean percentages table
    with st.expander("📈 Group Mean Percentages"):
        if is_threeway:
            comp_group_col_summary = '_factorial_group_'
            summary_df_tw = pd.concat([data[[results.factor_a_col, results.factor_b_col, results.factor_c_col]].reset_index(drop=True),
                                       percentages[selected_pct_cols].reset_index(drop=True)], axis=1)
            summary_df_tw[comp_group_col_summary] = (summary_df_tw[results.factor_a_col].astype(str) + ' - '
                                                     + summary_df_tw[results.factor_b_col].astype(str) + ' - '
                                                     + summary_df_tw[results.factor_c_col].astype(str))
            summary_data = []
            for pct_col in selected_pct_cols:
                ba_name = pct_col.replace('_pct', '')
                for group in summary_df_tw[comp_group_col_summary].unique():
                    group_data = summary_df_tw[summary_df_tw[comp_group_col_summary] == group][pct_col]
                    summary_data.append({
                        'Bile Acid': ba_name,
                        'Group': group,
                        'Mean %': f"{group_data.mean():.2f}",
                        'SD': f"{group_data.std():.2f}",
                        'Median %': f"{group_data.median():.2f}"
                    })
            if summary_data:
                pivot_df = pd.DataFrame(summary_data).pivot_table(
                    index='Bile Acid', columns='Group', values='Mean %', aggfunc='first')
                st.dataframe(pivot_df, width="stretch")
        elif is_twoway:
            comp_group_col_summary = '_factorial_group_'
            summary_df_tw = pd.concat([data[[results.factor_a_col, results.factor_b_col]].reset_index(drop=True),
                                       percentages[selected_pct_cols].reset_index(drop=True)], axis=1)
            summary_df_tw[comp_group_col_summary] = summary_df_tw[results.factor_a_col].astype(str) + ' - ' + summary_df_tw[results.factor_b_col].astype(str)
            summary_data = []
            for pct_col in selected_pct_cols:
                ba_name = pct_col.replace('_pct', '')
                for group in summary_df_tw[comp_group_col_summary].unique():
                    group_data = summary_df_tw[summary_df_tw[comp_group_col_summary] == group][pct_col]
                    summary_data.append({
                        'Bile Acid': ba_name,
                        'Group': group,
                        'Mean %': f"{group_data.mean():.2f}",
                        'SD': f"{group_data.std():.2f}",
                        'Median %': f"{group_data.median():.2f}"
                    })
            if summary_data:
                pivot_df = pd.DataFrame(summary_data).pivot_table(
                    index='Bile Acid', columns='Group', values='Mean %', aggfunc='first')
                st.dataframe(pivot_df, width="stretch")
        else:
            summary_data = []
            for pct_col in selected_pct_cols:
                ba_name = pct_col.replace('_pct', '')
                for group in plot_df[group_col].unique():
                    group_data = plot_df[plot_df[group_col] == group][pct_col]
                    summary_data.append({
                        'Bile Acid': ba_name,
                        'Group': group,
                        'Mean %': f"{group_data.mean():.2f}",
                        'SD': f"{group_data.std():.2f}",
                        'Median %': f"{group_data.median():.2f}"
                    })
            if summary_data:
                pivot_df = pd.DataFrame(summary_data).pivot_table(
                    index='Bile Acid', columns='Group', values='Mean %', aggfunc='first')
                st.dataframe(pivot_df, width="stretch")

    # Raw data view
    with st.expander("📋 View percentage data"):
        if is_threeway:
            view_df = pd.concat([data[[results.factor_a_col, results.factor_b_col, results.factor_c_col]].reset_index(drop=True),
                                percentages[selected_pct_cols].reset_index(drop=True)], axis=1)
            st.dataframe(view_df, width="stretch")
        elif is_twoway:
            view_df = pd.concat([data[[results.factor_a_col, results.factor_b_col]].reset_index(drop=True),
                                percentages[selected_pct_cols].reset_index(drop=True)], axis=1)
            st.dataframe(view_df, width="stretch")
        else:
            st.dataframe(plot_df, width="stretch")


def render_ratios_tab(processed, settings):
    """Render ratios tab with multi-panel display."""
    st.markdown("### Clinical Ratios")

    viz = BileAcidVisualizer(color_palette=settings['color_palette'], style=settings['plot_style'])
    group_col = processed.structure.group_col
    results = st.session_state.analysis_results
    is_threeway = results is not None and getattr(results, 'is_threeway', False)
    is_twoway = results is not None and results.is_twoway

    if not group_col:
        st.warning("No group column detected.")
        return

    mask = processed.sample_data[group_col].notna() & (processed.sample_data[group_col].astype(str).str.lower() != 'nan')
    data = processed.sample_data.loc[mask]

    if is_threeway:
        fa_col, fb_col, fc_col = results.factor_a_col, results.factor_b_col, results.factor_c_col
        combined = pd.concat([data[[fa_col, fb_col, fc_col]], processed.ratios.loc[data.index]], axis=1)
    elif is_twoway:
        fa_col = results.factor_a_col
        fb_col = results.factor_b_col
        combined = pd.concat([data[[fa_col, fb_col]], processed.ratios.loc[data.index]], axis=1)
    else:
        combined = pd.concat([data[[group_col]], processed.ratios.loc[data.index]], axis=1)

    available = [c for c in processed.ratios.columns if not combined[c].isna().all()]

    if not available:
        st.warning("No ratio data available.")
        return

    # Quick select options
    quick = st.segmented_control("Quick select",
                                  ["All ratios", "Significant only", "Key ratios", "Custom"],
                                  default="All ratios", key="ratio_quick")

    # Define key clinical ratios
    key_ratios = ['primary_to_secondary', 'glycine_to_taurine', 'conjugated_to_unconjugated',
                  '12alpha_to_non12alpha', 'primary_conjugated_to_secondary_conjugated',
                  'primary_unconjugated_to_secondary_unconjugated',
                  'CA_to_CDCA', 'TCA_to_GCA', 'GCDCA_to_TCDCA']

    if quick == "All ratios":
        selected_ratios = available
    elif quick == "Significant only":
        if is_threeway:
            selected_ratios = [r for r in available if r in results.threeway_ratios
                             and any(getattr(results.threeway_ratios[r].threeway_result, attr) < settings['alpha']
                                     for attr in ['factor_a_pvalue', 'factor_b_pvalue', 'factor_c_pvalue'])]
        elif is_twoway:
            selected_ratios = [r for r in available if r in results.twoway_ratios
                             and (results.twoway_ratios[r].twoway_result.factor_a_pvalue < settings['alpha']
                                  or results.twoway_ratios[r].twoway_result.factor_b_pvalue < settings['alpha']
                                  or results.twoway_ratios[r].twoway_result.interaction_pvalue < settings['alpha'])]
        else:
            selected_ratios = [r for r in available if r in results.ratios_results
                             and results.ratios_results[r].main_test.significant]
        if not selected_ratios:
            st.info("No significant ratio differences found. Showing all ratios.")
            selected_ratios = available
    elif quick == "Key ratios":
        selected_ratios = [r for r in key_ratios if r in available]
        if not selected_ratios:
            selected_ratios = available[:6]
    else:
        selected_ratios = st.multiselect("Select ratios", available,
                                         default=available[:6] if len(available) >= 6 else available,
                                         key="ratio_custom")

    if not selected_ratios:
        st.warning("No ratios selected.")
        return

    # Show significant findings
    if is_threeway:
        sig_ratios = [r for r in selected_ratios if r in results.threeway_ratios
                     and any(getattr(results.threeway_ratios[r].threeway_result, attr) < settings['alpha']
                             for attr in ['factor_a_pvalue', 'factor_b_pvalue', 'factor_c_pvalue'])]
    elif is_twoway:
        sig_ratios = [r for r in selected_ratios if r in results.twoway_ratios
                     and (results.twoway_ratios[r].twoway_result.factor_a_pvalue < settings['alpha']
                          or results.twoway_ratios[r].twoway_result.factor_b_pvalue < settings['alpha']
                          or results.twoway_ratios[r].twoway_result.interaction_pvalue < settings['alpha'])]
    else:
        sig_ratios = [r for r in selected_ratios if r in results.ratios_results
                     and results.ratios_results[r].main_test.significant]
    if sig_ratios:
        st.success(f"**Significant differences:** {', '.join(sig_ratios)}")

    log_scale = st.checkbox("Log₁₀ scale", key="ratio_log")

    if is_threeway:
        fa_name, fb_name, fc_name = results.factor_a_name, results.factor_b_name, results.factor_c_name
        st.markdown(f"**Three-way ANOVA**: {fa_name} x {fb_name} x {fc_name}")

        tw_stats = {r: results.threeway_ratios.get(r) for r in selected_ratios
                   if r in results.threeway_ratios}

        fig = viz.plot_threeway_multi_panel(
            data=combined, value_cols=selected_ratios,
            factor_a_col=fa_col, factor_b_col=fb_col, factor_c_col=fc_col,
            threeway_results=tw_stats, ncols=3,
            factor_a_name=fa_name, factor_b_name=fb_name, factor_c_name=fc_name,
            ylabel='Ratio', show_points=settings['show_points'],
            plot_type=settings['plot_type']
        )
        st.pyplot(fig)
        store_figure(fig, 'ratios_threeway')
        plt.close(fig)

        with st.expander("Interaction Plots"):
            fig_int = viz.plot_threeway_interaction_multi_panel(
                data=combined, value_cols=selected_ratios,
                factor_a_col=fa_col, factor_b_col=fb_col, factor_c_col=fc_col,
                threeway_results=tw_stats, ncols=3,
                factor_a_name=fa_name, factor_b_name=fb_name, factor_c_name=fc_name,
                ylabel='Ratio'
            )
            st.pyplot(fig_int)
            store_figure(fig_int, 'ratios_threeway_interaction')
            plt.close(fig_int)

        with st.expander("Three-Way ANOVA Summary"):
            if tw_stats:
                st.dataframe(get_threeway_differences_summary(tw_stats), hide_index=True)
    elif is_twoway:
        fa_name = results.factor_a_name
        fb_name = results.factor_b_name
        st.markdown(f"**Two-way ANOVA**: {fa_name} x {fb_name}")

        tw_stats = {r: results.twoway_ratios.get(r) for r in selected_ratios
                   if r in results.twoway_ratios}

        fig = viz.plot_twoway_multi_panel(
            data=combined, value_cols=selected_ratios,
            factor_a_col=fa_col, factor_b_col=fb_col,
            twoway_results=tw_stats, ncols=3,
            factor_a_name=fa_name, factor_b_name=fb_name,
            ylabel='Ratio', show_points=settings['show_points'],
            plot_type=settings['plot_type']
        )
        st.pyplot(fig)
        store_figure(fig, 'ratios_twoway')
        plt.close(fig)

        with st.expander("Interaction Plots"):
            fig_int = viz.plot_twoway_interaction_multi_panel(
                data=combined, value_cols=selected_ratios,
                factor_a_col=fa_col, factor_b_col=fb_col,
                twoway_results=tw_stats, ncols=3,
                factor_a_name=fa_name, factor_b_name=fb_name,
                ylabel='Ratio'
            )
            st.pyplot(fig_int)
            store_figure(fig_int, 'ratios_interaction')
            plt.close(fig_int)

        with st.expander("Two-Way ANOVA Summary"):
            if tw_stats:
                st.dataframe(get_twoway_differences_summary(tw_stats), hide_index=True)
    else:
        # Get stats for selected ratios
        stats_dict = {r: results.ratios_results.get(r) for r in selected_ratios if r in results.ratios_results}

        # Multi-panel figure
        fig = viz.plot_multi_panel_groups_with_stats(
            combined, selected_ratios, group_col, stats_dict,
            ncols=3, plot_type=settings['plot_type'], log_scale=log_scale,
            show_points=settings['show_points'], ylabel='Ratio'
        )
        st.pyplot(fig)
        store_figure(fig, f'ratios{"_log" if log_scale else ""}')
        plt.close(fig)

        # Generate other version for export
        fig_other = viz.plot_multi_panel_groups_with_stats(
            combined, selected_ratios, group_col, stats_dict,
            ncols=3, plot_type=settings['plot_type'], log_scale=not log_scale,
            show_points=settings['show_points'], ylabel='Ratio'
        )
        store_figure(fig_other, f'ratios{"_log" if not log_scale else ""}')
        plt.close(fig_other)

        # Statistical summary
        with st.expander("📊 Statistical Summary"):
            rows = []
            for r in selected_ratios:
                res = results.ratios_results.get(r)
                if res:
                    n_sig = 0
                    if res.posthoc_test and res.posthoc_test.pairwise_results is not None:
                        n_sig = res.posthoc_test.pairwise_results['significant'].sum()
                    rows.append({
                        'Ratio': r,
                        'Test': res.main_test.test_type.value,
                        'P-value': f"{res.main_test.pvalue:.4f}",
                        'Significant': '✓' if res.main_test.significant else '',
                        'Effect Size': f"{res.main_test.effect_size:.3f}" if res.main_test.effect_size else 'N/A',
                        'Sig. Pairs': n_sig,
                        'APA': format_apa_statistics(res)
                    })
            if rows:
                st.dataframe(pd.DataFrame(rows), hide_index=True)

        # Group means table
        with st.expander("📈 Group Mean Ratios"):
            summary_data = []
            for ratio in selected_ratios:
                for group in combined[group_col].unique():
                    group_data = combined[combined[group_col] == group][ratio]
                    summary_data.append({
                        'Ratio': ratio,
                        'Group': group,
                        'Mean': f"{group_data.mean():.3f}",
                        'SD': f"{group_data.std():.3f}",
                        'Median': f"{group_data.median():.3f}"
                    })
            if summary_data:
                summary_df = pd.DataFrame(summary_data)
                pivot_df = summary_df.pivot_table(index='Ratio', columns='Group', values='Mean', aggfunc='first')
                st.dataframe(pivot_df, width="stretch")

    # Raw data
    with st.expander("📋 View ratio data"):
        if is_threeway:
            display_cols = [fa_col, fb_col, fc_col] + selected_ratios
        elif is_twoway:
            display_cols = [fa_col, fb_col] + selected_ratios
        else:
            display_cols = [group_col] + selected_ratios
        st.dataframe(combined[display_cols], width="stretch")


def _count_sig_twoway(twoway_dict, alpha=0.05):
    """Count significant results in a two-way ANOVA results dictionary."""
    count = 0
    for res in twoway_dict.values():
        if (res.twoway_result.factor_a_pvalue < alpha
                or res.twoway_result.factor_b_pvalue < alpha
                or res.twoway_result.interaction_pvalue < alpha):
            count += 1
    return count


def _count_sig_threeway(threeway_dict, alpha=0.05):
    """Count significant results in a three-way ANOVA results dictionary."""
    count = 0
    for res in threeway_dict.values():
        if any(getattr(res.threeway_result, attr) < alpha
               for attr in ['factor_a_pvalue', 'factor_b_pvalue', 'factor_c_pvalue']):
            count += 1
    return count


def render_statistics_tab(processed, settings):
    """Statistics summary tab."""
    st.markdown("### Statistics Summary")
    results = st.session_state.analysis_results
    is_threeway = results is not None and getattr(results, 'is_threeway', False)
    is_twoway = results is not None and results.is_twoway
    alpha = settings['alpha']

    if is_threeway:
        sig_t = _count_sig_threeway(results.threeway_totals, alpha)
        sig_b = _count_sig_threeway(results.threeway_individual_ba, alpha)
        sig_p = _count_sig_threeway(results.threeway_percentages, alpha)
        sig_r = _count_sig_threeway(results.threeway_ratios, alpha)

        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Sig. Totals", f"{sig_t}/{len(results.threeway_totals)}")
        c2.metric("Sig. BAs", f"{sig_b}/{len(results.threeway_individual_ba)}")
        c3.metric("Sig. Percentages", f"{sig_p}/{len(results.threeway_percentages)}")
        c4.metric("Sig. Ratios", f"{sig_r}/{len(results.threeway_ratios)}")

        st.markdown(f"**Design**: {results.factor_a_name} x {results.factor_b_name} x {results.factor_c_name} (Three-Way ANOVA)")

        st.markdown("#### Totals")
        if results.threeway_totals:
            st.dataframe(get_threeway_differences_summary(results.threeway_totals), hide_index=True)

        st.markdown("#### Individual BAs")
        if results.threeway_individual_ba:
            sig_ba_tw = {k: v for k, v in results.threeway_individual_ba.items()
                        if any(getattr(v.threeway_result, attr) < alpha
                               for attr in ['factor_a_pvalue', 'factor_b_pvalue', 'factor_c_pvalue'])}
            if sig_ba_tw:
                st.dataframe(get_threeway_differences_summary(sig_ba_tw), hide_index=True)
            else:
                st.info("No significant individual BA differences.")

        st.markdown("#### Percentages")
        if results.threeway_percentages:
            sig_pct_tw = {k.replace('_pct', ''): v for k, v in results.threeway_percentages.items()
                         if any(getattr(v.threeway_result, attr) < alpha
                                for attr in ['factor_a_pvalue', 'factor_b_pvalue', 'factor_c_pvalue'])}
            if sig_pct_tw:
                st.dataframe(get_threeway_differences_summary(sig_pct_tw), hide_index=True)
            else:
                st.info("No significant percentage differences.")

        st.markdown("#### Ratios")
        if results.threeway_ratios:
            sig_rat_tw = {k: v for k, v in results.threeway_ratios.items()
                         if any(getattr(v.threeway_result, attr) < alpha
                                for attr in ['factor_a_pvalue', 'factor_b_pvalue', 'factor_c_pvalue'])}
            if sig_rat_tw:
                st.dataframe(get_threeway_differences_summary(sig_rat_tw), hide_index=True)
            else:
                st.info("No significant ratio differences.")

        # APA report
        with st.expander("📝 APA-Formatted Report"):
            for category, tw_dict, label in [
                ("Totals", results.threeway_totals, "totals"),
                ("Individual BAs", results.threeway_individual_ba, "individual"),
                ("Ratios", results.threeway_ratios, "ratios"),
            ]:
                if tw_dict:
                    st.markdown(f"**{category}:**")
                    for name, res in tw_dict.items():
                        apa = format_threeway_apa(res)
                        if apa:
                            st.markdown(f"*{name}*: {apa}")
    elif is_twoway:
        sig_t = _count_sig_twoway(results.twoway_totals, alpha)
        sig_b = _count_sig_twoway(results.twoway_individual_ba, alpha)
        sig_p = _count_sig_twoway(results.twoway_percentages, alpha)
        sig_r = _count_sig_twoway(results.twoway_ratios, alpha)

        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Sig. Totals", f"{sig_t}/{len(results.twoway_totals)}")
        c2.metric("Sig. BAs", f"{sig_b}/{len(results.twoway_individual_ba)}")
        c3.metric("Sig. Percentages", f"{sig_p}/{len(results.twoway_percentages)}")
        c4.metric("Sig. Ratios", f"{sig_r}/{len(results.twoway_ratios)}")

        st.markdown(f"**Design**: {results.factor_a_name} x {results.factor_b_name} (Two-Way ANOVA)")

        st.markdown("#### Totals")
        if results.twoway_totals:
            st.dataframe(get_twoway_differences_summary(results.twoway_totals), hide_index=True)

        st.markdown("#### Individual BAs")
        if results.twoway_individual_ba:
            sig_ba_tw = {k: v for k, v in results.twoway_individual_ba.items()
                        if (v.twoway_result.factor_a_pvalue < alpha
                            or v.twoway_result.factor_b_pvalue < alpha
                            or v.twoway_result.interaction_pvalue < alpha)}
            if sig_ba_tw:
                st.dataframe(get_twoway_differences_summary(sig_ba_tw), hide_index=True)
            else:
                st.info("No significant individual BA differences.")

        st.markdown("#### Percentages")
        if results.twoway_percentages:
            sig_pct_tw = {k.replace('_pct', ''): v for k, v in results.twoway_percentages.items()
                         if (v.twoway_result.factor_a_pvalue < alpha
                             or v.twoway_result.factor_b_pvalue < alpha
                             or v.twoway_result.interaction_pvalue < alpha)}
            if sig_pct_tw:
                st.dataframe(get_twoway_differences_summary(sig_pct_tw), hide_index=True)
            else:
                st.info("No significant percentage differences.")

        st.markdown("#### Ratios")
        if results.twoway_ratios:
            sig_rat_tw = {k: v for k, v in results.twoway_ratios.items()
                         if (v.twoway_result.factor_a_pvalue < alpha
                             or v.twoway_result.factor_b_pvalue < alpha
                             or v.twoway_result.interaction_pvalue < alpha)}
            if sig_rat_tw:
                st.dataframe(get_twoway_differences_summary(sig_rat_tw), hide_index=True)
            else:
                st.info("No significant ratio differences.")

        # APA report
        with st.expander("📝 APA-Formatted Report"):
            for category, tw_dict, label in [
                ("Totals", results.twoway_totals, "totals"),
                ("Individual BAs", results.twoway_individual_ba, "individual"),
                ("Ratios", results.twoway_ratios, "ratios"),
            ]:
                if tw_dict:
                    st.markdown(f"**{category}:**")
                    for name, res in tw_dict.items():
                        apa = format_twoway_apa(res)
                        if apa:
                            st.markdown(f"*{name}*: {apa}")
    else:
        sig_t = sum(1 for r in results.totals_results.values() if r.main_test.significant)
        sig_b = sum(1 for r in results.individual_ba_results.values() if r.main_test.significant)
        sig_p = sum(1 for r in results.percentages_results.values() if r.main_test.significant)
        sig_r = sum(1 for r in results.ratios_results.values() if r.main_test.significant)

        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Sig. Totals", f"{sig_t}/{len(results.totals_results)}")
        c2.metric("Sig. BAs", f"{sig_b}/{len(results.individual_ba_results)}")
        c3.metric("Sig. Percentages", f"{sig_p}/{len(results.percentages_results)}")
        c4.metric("Sig. Ratios", f"{sig_r}/{len(results.ratios_results)}")

        st.markdown("#### Totals")
        st.dataframe(get_significant_differences_summary(results.totals_results), hide_index=True)

        st.markdown("#### Individual BAs (Significant)")
        sig_ba = {k: v for k, v in results.individual_ba_results.items() if v.main_test.significant}
        if sig_ba:
            st.dataframe(get_significant_differences_summary(sig_ba), hide_index=True)
        else:
            st.info("No significant individual BA differences.")

        st.markdown("#### Percentages (Significant)")
        sig_pct = {k: v for k, v in results.percentages_results.items() if v.main_test.significant}
        if sig_pct:
            sig_pct_display = {k.replace('_pct', ''): v for k, v in sig_pct.items()}
            st.dataframe(get_significant_differences_summary(sig_pct_display), hide_index=True)
        else:
            st.info("No significant percentage differences.")


@st.fragment
def render_export_tab(processed, settings):
    """Export tab."""
    st.markdown("### Export Results")
    results = st.session_state.analysis_results
    report_gen = st.session_state.report_generator
    
    col1, col2 = st.columns(2)
    with col1:
        st.markdown("#### Individual Downloads")
    
    # Get metadata columns (Sample_ID and Group)
        id_cols = []
        if processed.structure.sample_id_col and processed.structure.sample_id_col in processed.sample_data.columns:
            id_cols.append(processed.structure.sample_id_col)
        if processed.structure.group_col and processed.structure.group_col in processed.sample_data.columns:
            id_cols.append(processed.structure.group_col)
    
        metadata = processed.sample_data[id_cols] if id_cols else pd.DataFrame(index=processed.sample_data.index)
    
        # Export each dataset with metadata
        for name, df in [("Concentrations", processed.concentrations),
                     ("Totals", processed.totals),
                     ("Ratios", processed.ratios),
                     ("Percentages", processed.percentages)]:
            export_df = pd.concat([metadata, df], axis=1)
            st.download_button(
                f"📥 {name} (Excel)",
                df_to_excel_bytes(export_df),
                f"bile_acid_{name.lower()}.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"download_{name.lower()}"
            )
        
        # LOD-highlighted Excel export
        st.markdown("---")
        lod_excel = create_full_data_excel_with_highlighting(processed, settings['lod_handling'])
        lod_label = {"half_lod": "LOD/2", "lod": "LOD value", "zero": "Zero", "drop": "NaN"}
        st.download_button(
            "📥 All Data (Excel) - LOD highlighted",
            lod_excel,
            f"bile_acid_data_lod_highlighted_{datetime.now():%Y%m%d}.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_lod_excel"
        )
        st.caption(f"Yellow cells = values replaced (below LOD → {lod_label[settings['lod_handling']]})")
    
    with col2:
        st.markdown("#### Complete Reports")
        if report_gen:
            buf = BytesIO()
            report_gen.save_excel_report(buf)
            buf.seek(0)
            st.download_button("📥 Statistical Report (Excel)", buf.getvalue(),
                              f"statistical_report_{datetime.now():%Y%m%d}.xlsx",
                              "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                              key="download_excel_report")
        
        # Generate ZIP on button click, then show download button
        if st.button("📦 Generate Complete Package (ZIP)", key="generate_zip"):
            with st.status("Generating export package...", expanded=True) as status:
                st.write("Rendering all figure variations...")
                all_figures = generate_all_export_figures(processed, results, settings)
                combined_figures = {**st.session_state.figures, **all_figures}
                st.write("Packaging ZIP archive...")
                st.session_state.zip_data = create_results_zip(processed, results, combined_figures, report_gen, settings)
                st.session_state.zip_figure_count = len(combined_figures)
                status.update(label="Package ready!", state="complete", expanded=False)

        if st.session_state.zip_data is not None:
            st.download_button("📥 Download Complete Package (ZIP)", st.session_state.zip_data,
                              f"bile_acid_analysis_{datetime.now():%Y%m%d_%H%M}.zip",
                              "application/zip",
                              key="download_zip")
            st.caption(f"📊 Package includes {st.session_state.zip_figure_count} figures covering all analysis options")
    
    st.markdown("---")



def render_heatmap_tab(processed, settings):
    """Render heatmap tab - faceted heatmap of bile acid expression by category."""
    st.markdown("### Faceted Heatmap")
    st.caption("Visualize bile acid expression patterns across samples, grouped by category")

    viz = BileAcidVisualizer(color_palette=settings['color_palette'], style=settings['plot_style'])
    group_col = processed.structure.group_col

    if not group_col:
        st.warning("No group column detected.")
        return

    available_bas = set(processed.structure.bile_acid_cols)

    # Build available categories (only those with detected BAs and a totals column)
    available_categories = {}
    for cat_key, ba_list in ANALYSIS_GROUPS.items():
        present_bas = [ba for ba in ba_list if ba in available_bas]
        if present_bas and cat_key in processed.totals.columns:
            display = ANALYSIS_GROUP_DISPLAY_NAMES.get(cat_key, cat_key)
            available_categories[display] = cat_key

    if not available_categories:
        st.warning("No categories available.")
        return

    # Category selection
    sorted_display_names = sorted(available_categories.keys())
    default_selection = [n for n in ["Primary", "Secondary", "Oxidized (Keto)", "Epimerized (Iso)"]
                        if n in sorted_display_names]
    if not default_selection:
        default_selection = sorted_display_names[:3]

    selected_display = st.multiselect(
        "Select categories",
        sorted_display_names,
        default=default_selection,
        key="heatmap_categories"
    )

    if not selected_display:
        st.info("Select one or more categories to generate a heatmap.")
        return

    # Options
    col1, col2, col3 = st.columns(3)
    with col1:
        norm_mode = st.selectbox("Normalization", ["Z-score", "Row %", "None"],
                                 index=0, key="heatmap_norm")
        norm_map = {"Z-score": "zscore", "Row %": "row", "None": "none"}
        normalization = norm_map[norm_mode]
    with col2:
        cmap = st.selectbox("Color map", ["RdBu_r", "coolwarm", "viridis", "YlOrRd", "PiYG"],
                            index=0, key="heatmap_cmap")
    with col3:
        st.write("")  # spacer

    # Build selected_categories dict: {cat_key: [BA1, BA2, ...]}
    selected_cats = {}
    for display in selected_display:
        cat_key = available_categories[display]
        present_bas = [ba for ba in ANALYSIS_GROUPS[cat_key] if ba in available_bas]
        selected_cats[cat_key] = present_bas

    # Generate button
    if st.button("Generate Heatmap", type="primary", key="heatmap_generate"):
        with st.spinner("Generating heatmap..."):
            hmask = processed.sample_data[group_col].notna() & (processed.sample_data[group_col].astype(str).str.lower() != 'nan')
            data = processed.sample_data.loc[hmask]

            # Build factor_cols: use detected factors, fallback to group_col
            if processed.structure.factors:
                fc = processed.structure.factors
            else:
                fc = {group_col: group_col}

            fig = viz.plot_faceted_heatmap(
                sample_data=data,
                totals=processed.totals.loc[data.index],
                group_col=group_col,
                selected_categories=selected_cats,
                category_display_names=ANALYSIS_GROUP_DISPLAY_NAMES,
                normalization=normalization,
                cmap=cmap,
                sample_id_col=processed.structure.sample_id_col,
                factor_cols=fc,
            )
            st.session_state['heatmap_fig'] = fig_to_bytes(fig)
            store_figure(fig, 'faceted_heatmap')
            plt.close(fig)

    # Display cached heatmap
    if 'heatmap_fig' in st.session_state and st.session_state['heatmap_fig'] is not None:
        st.image(st.session_state['heatmap_fig'], use_container_width=True)


def main():
    """Main entry point."""
    init_session_state()
    settings = render_sidebar()
    
    st.markdown("# 🧬 Bile Acid Analysis Pipeline")
    uploaded = st.file_uploader("Upload Excel/ODS file", ['xlsx', 'xls', 'ods'])
    
    if uploaded:
        # Check if file changed
        file_changed = st.session_state.last_file != uploaded.name
        if file_changed:
            st.session_state.stats_computed = False
            st.session_state.stats_sections_done = set()
            st.session_state.figures = {}
            st.session_state.report_generator = None
            st.session_state.zip_data = None
            st.session_state.zip_figure_count = 0
            st.session_state.last_file = uploaded.name
            st.session_state.last_settings = None  # Reset settings tracking for new file
            st.session_state.selected_sheet = None  # Reset sheet selection for new file

        # Write temp file for sheet inspection and processing
        with tempfile.NamedTemporaryFile(delete=False, suffix=Path(uploaded.name).suffix) as tmp:
            tmp.write(uploaded.getvalue())
            tmp_path = tmp.name

        # Sheet selection dropdown
        processor = BileAcidDataProcessor(lod_handling=settings['lod_handling'])
        try:
            selectable_sheets, auto_detected_sheet = processor.get_selectable_sheets(tmp_path)
        except Exception as e:
            st.error(f"Error reading file: {e}")
            return

        if len(selectable_sheets) > 1:
            # Determine default: use previous selection if still valid, otherwise auto-detected
            if st.session_state.selected_sheet in selectable_sheets:
                default_idx = selectable_sheets.index(st.session_state.selected_sheet)
            else:
                default_idx = selectable_sheets.index(auto_detected_sheet)

            selected_sheet = st.selectbox(
                "Select data sheet",
                selectable_sheets,
                index=default_idx,
                help="Choose which sheet contains your sample data. The LC-MS data sheet (standard curves) is excluded."
            )
        else:
            selected_sheet = selectable_sheets[0] if selectable_sheets else None

        # Detect sheet change
        sheet_changed = (st.session_state.selected_sheet != selected_sheet)
        if sheet_changed:
            st.session_state.selected_sheet = selected_sheet
            st.session_state.stats_computed = False
            st.session_state.stats_sections_done = set()
            st.session_state.report_generator = None
            st.session_state.figures = {}
            st.session_state.zip_data = None
            st.session_state.zip_figure_count = 0

        # Check if settings changed (LOD handling, alpha, etc.)
        settings_changed = check_settings_changed(settings)
        if settings_changed:
            st.toast("Settings changed — reprocessing data...")

        # Reprocess data if file, sheet, or settings changed
        if file_changed or sheet_changed or settings_changed or st.session_state.processed_data is None:
            with st.status("Processing data...", expanded=True) as status:
                st.write(f"Reading sheet: **{selected_sheet}**...")

                try:
                    st.write("Detecting structure & applying LOD handling...")
                    processed = processor.load_and_process(tmp_path, sheet_name=selected_sheet)
                    st.session_state.processed_data = processed
                    status.update(label="Data processed!", state="complete", expanded=False)
                except Exception as e:
                    status.update(label="Processing failed", state="error")
                    st.error(f"Error: {e}")
                    return
    
    if st.session_state.processed_data:
        processed = st.session_state.processed_data
        
        quality = validate_data_quality(processed)
        c1, c2, c3 = st.columns(3)
        c1.metric("Samples", quality['n_samples'])
        c2.metric("Bile Acids", quality['n_bile_acids_detected'])
        c3.metric("Groups", quality.get('n_groups', 'N/A'))
        
        # Show LOD info
        lod_source_label = "Auto-detected from standards" if quality.get('lod_source') == 'standards' else "Default (no standards found)"
        lod_display = {"half_lod": "LOD/2", "lod": "LOD value", "zero": "Zero", "drop": "NaN"}
        st.caption(f"📊 LOD: **{lod_source_label}** (below-LOD → {lod_display[settings['lod_handling']]}) | α = **{settings['alpha']}**")

        # Show factor detection info
        if processed.structure.n_factors >= 3:
            factor_names = list(processed.structure.factors.keys())
            source_label = {"metadata_sheet": "metadata sheet", "prefix_columns": "Factor_ columns"}.get(
                processed.structure.factor_source, "unknown")
            st.info(f"🔬 **Three-way design detected** ({source_label}): **{factor_names[0]}** × **{factor_names[1]}** × **{factor_names[2]}**")
        elif processed.structure.n_factors >= 2:
            factor_names = list(processed.structure.factors.keys())
            source_label = {"metadata_sheet": "metadata sheet", "prefix_columns": "Factor_ columns"}.get(
                processed.structure.factor_source, "unknown")
            st.info(f"🔬 **Two-way design detected** ({source_label}): **{factor_names[0]}** × **{factor_names[1]}**")
        
        # Show per-analyte LOD details
        if quality.get('analyte_lods'):
            with st.expander("🔬 Per-Analyte LOD Details"):
                lod_rows = []
                for ba in processed.structure.bile_acid_cols:
                    lod_val = quality['analyte_lods'].get(ba, 'N/A')
                    count = quality.get('analyte_lod_counts', {}).get(ba, 0)
                    n_samples = quality['n_samples']
                    pct = f"{count/n_samples*100:.0f}%" if n_samples > 0 else "0%"
                    lod_rows.append({
                        'Analyte': ba,
                        'LOD (nM)': lod_val if isinstance(lod_val, str) else f"{lod_val:.1f}",
                        'Below LOD': count,
                        '% Below': pct
                    })
                if lod_rows:
                    lod_df = pd.DataFrame(lod_rows)
                    # Highlight high replacement rates
                    st.dataframe(lod_df, width="stretch", hide_index=True)
        
        n_factors = getattr(processed.structure, 'n_factors', 0)
        is_threeway = n_factors >= 3

        if is_threeway:
            # Three-way: use a selector so only ONE section computes at a time
            # (st.tabs renders all content on every run, which OOMs on Cloud)
            tab_options = ["📈 Concentrations", "📊 Totals", "📉 Percentages", "🔢 Ratios",
                           "🗺️ Heatmap", "📊 Statistics", "💾 Export"]
            active_tab = st.segmented_control("Section", tab_options, default="📈 Concentrations")

            # Map tab to stats section(s) needed
            tab_sections = {
                "📈 Concentrations": ['individual_ba'],
                "📊 Totals": ['totals'],
                "📉 Percentages": ['percentages'],
                "🔢 Ratios": ['ratios'],
                "🗺️ Heatmap": [],
                "📊 Statistics": ['individual_ba', 'totals', 'ratios', 'percentages', 'categories'],
                "💾 Export": ['individual_ba', 'totals', 'ratios', 'percentages', 'categories'],
            }
            needed = tab_sections.get(active_tab, [])
            missing = [s for s in needed if s not in st.session_state.stats_sections_done]
            if missing:
                label = f"Computing {', '.join(missing)} statistics..."
                with st.status(label, expanded=False) as status:
                    for sec in missing:
                        ensure_section_computed(processed, settings, sec)
                    status.update(label="Statistics ready", state="complete")

            # Always free matplotlib/gc memory before rendering a new tab
            plt.close('all')
            gc.collect()

            if active_tab == "📈 Concentrations":
                render_concentrations_tab(processed, settings)
            elif active_tab == "📊 Totals":
                render_totals_tab(processed, settings)
            elif active_tab == "📉 Percentages":
                render_percentages_tab(processed, settings)
            elif active_tab == "🔢 Ratios":
                render_ratios_tab(processed, settings)
            elif active_tab == "🗺️ Heatmap":
                render_heatmap_tab(processed, settings)
            elif active_tab == "📊 Statistics":
                render_statistics_tab(processed, settings)
            elif active_tab == "💾 Export":
                render_export_tab(processed, settings)
        else:
            # One-way / two-way: compute all stats at once (fits in memory), use tabs
            with st.status("Computing statistics...", expanded=False) as status:
                compute_all_statistics(processed, settings)
                status.update(label="Statistics ready", state="complete")

            tabs = st.tabs(["📈 Concentrations", "📊 Totals", "📉 Percentages", "🔢 Ratios",
                            "🗺️ Heatmap", "📊 Statistics", "💾 Export"])
            with tabs[0]: render_concentrations_tab(processed, settings)
            with tabs[1]: render_totals_tab(processed, settings)
            with tabs[2]: render_percentages_tab(processed, settings)
            with tabs[3]: render_ratios_tab(processed, settings)
            with tabs[4]: render_heatmap_tab(processed, settings)
            with tabs[5]: render_statistics_tab(processed, settings)
            with tabs[6]: render_export_tab(processed, settings)
    
    else:
        # Show expected input format when no data is loaded
        st.markdown("---")
        st.markdown("## 📋 Expected Input Format")
        
        st.markdown("### LC-MS sheet")
        st.markdown("""
        - **Rows:** First few are Std curves, followed by samples
        - **Columns:** "Data Filename" (Std curve and sample names), followed by sphingolipid species
        """)


        st.markdown("**Example:**")
        
        # Create example dataframe
        example_lcms_df = pd.DataFrame({
            'Data Filename': ['Std 1 ngmL', 'Std 3  ngmL', 'Std 10 ngmL'],
            'CA': [1.3, 3.0, 9.9],
            'TCA': [1.1, 2.9, 10.1],
            'GCA': ['----', 3.1, 10.0],
            'TCDCA': ['----', '----', 10.2],
            '...': ['...', '...', '...']
        })
        st.dataframe(example_lcms_df.astype(str), hide_index=True, width="content")

        st.markdown("### Sample sheet")
        st.markdown("""
        - **Rows:** Samples
        - **Columns:** Sphingolipid species (matching panel names)
        - **First column(s):** Sample ID, Group/Type
        - **Values:** Concentrations (typically ng/mL)
        - **Below LOD:** Can be "-----", "LOD", "BLQ", "ND", etc.
        """)
        
        st.markdown("**Example:**")
        
        # Create example dataframe
        example_df = pd.DataFrame({
            'Type': ['Aged-1', 'Aged-2', 'Young-1'],
            'Sample_ID': ['S001', 'S002', 'S003'],
            'CA': [125.4, 142.8, 98.6],
            'TCA': [312.5, 287.9, '-----'],
            'GCA': [1520.3, 1380.7, 1245.2],
            'TCDCA': [15.2, 18.4, 12.8],
            '...': ['...', '...', '...']
        })
        st.dataframe(example_df.astype(str), hide_index=True, width="content")
        
        st.markdown("""**For Multiple Independent Variables need to add "Factor_" in front of it""")
        st.markdown("**Example:**")
        # Create example dataframe
        example_df = pd.DataFrame({
            'Factor_Age': ['Aged', 'Aged', 'Young'],
            'Factor_Sex': ['Male', 'Female', 'Female'],
            'Sample_ID': ['S001', 'S002', 'S003'],
            'CA': [125.4, 142.8, 98.6],
            'TCA': [312.5, 287.9, '-----'],
            'GCA': [1520.3, 1380.7, 1245.2],
            'TCDCA': [15.2, 18.4, 12.8],
            '...': ['...', '...', '...']
        })
        st.dataframe(example_df.astype(str), hide_index=True, width='content')

        st.info("💡 The pipeline auto-detects bile acid columns and group assignments from your data.")


if __name__ == "__main__":
    main()
